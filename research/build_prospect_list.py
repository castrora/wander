"""
Pierway — Port of Seattle Prospect List Builder v4
Generates research/port-of-seattle-prospects.xlsx

Sheets:
  1. Priority Contacts  — Port / Tourism / other high-leverage orgs
  2. Cruise Lines       — named contacts, emails, phones, pitch, tracking
  3. Hotels             — named contacts, emails, phones, pitch, tracking
  4. Other Targets      — ferries, attractions, tour operators, airports
  5. How To Use
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colours ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
TEAL       = "1A7A6E"
PURPLE     = "5C3D8F"
ORANGE     = "C55A11"
LIGHT_BLUE = "D6E4F0"
LIGHT_TEAL = "D6EFEC"
LIGHT_PURP = "EAE0F5"
LIGHT_ORG  = "FCE4D6"
WHITE      = "FFFFFF"
LIGHT_GREY = "F2F2F2"
YELLOW_BG  = "FFFACD"
GREEN_BG   = "E2EFDA"
GREEN_FG   = "375623"

def fill(h): return PatternFill("solid", fgColor=h)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row_num, bg, fg=WHITE):
    for cell in ws[row_num]:
        cell.font      = Font(bold=True, color=fg, name="Calibri", size=10)
        cell.fill      = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr()
    ws.row_dimensions[row_num].height = 40

def row_style(ws, r, bg=WHITE):
    for cell in ws[r]:
        cell.fill      = fill(bg)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = bdr()
        cell.font      = Font(name="Calibri", size=10)
    ws.row_dimensions[r].height = 48

def section_hdr(ws, r, bg, label, ncols):
    ws.append([""] * ncols)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=f"  {label}")
    c.font      = Font(bold=True, color=WHITE, name="Calibri", size=11)
    c.fill      = fill(bg)
    c.alignment = Alignment(vertical="center")
    c.border    = bdr()
    ws.row_dimensions[r].height = 22

def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v

# Tracking column headers (reused across sheets)
TRACK_HDRS = ["Outreach Sent?\n(Date or Y/N)",
              "Response?\n(Date or Y/N)",
              "Follow-up\nNeeded?",
              "Meeting\nScheduled?",
              "Status\nNotes"]

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — PRIORITY CONTACTS
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Priority Contacts"

H1 = (["Organization", "Category", "Contact Name", "Title",
        "Email", "Phone", "Targeted Pitch (QR / Pierway)",
        "LinkedIn / Source"] + TRACK_HDRS)
ws1.append(H1)
hdr(ws1, 1, DARK_BLUE)
NC1 = len(H1)

BLANK5 = ["", "", "", "", ""]

def add_sec(ws, r, bg, label, nc):
    section_hdr(ws, r, bg, label, nc)
    return r + 1

def add_row(ws, r, data, bg):
    ws.append(data + BLANK5)
    row_style(ws, r, bg)
    return r + 1

r = 2
r = add_sec(ws1, r, TEAL, "PORT OF SEATTLE", NC1)
port_rows = [
    ["Port of Seattle", "Port — Cruise BD",
     "Marie Ellingson", "Manager, Cruise Services & Business Development",
     "mellingson@portseattle.org", "(206) 787-3529",
     "Hi Marie — I'm building Pierway, an AI walking platform for cruise passengers in Seattle. I'd love to understand the vendor/partner process for QR code placement at Pier 91 and Pier 66. Could we set up a 20-minute call?",
     "portseattle.org/contacts/cruise-terminal-manager"],
    ["Port of Seattle", "Port — Public Affairs",
     "Rosie Courtney", "Sr. Manager, Cruise Public Affairs & Community Engagement",
     "", "",
     "Hi Rosie — Pierway is a new visitor experience platform launching in Seattle for cruise passengers. We'd love to explore whether our AI walking routes fit the Port's community engagement goals.",
     '"Port of Seattle" "Rosie Courtney"'],
]
for row_data in port_rows:
    r = add_row(ws1, r, row_data, LIGHT_TEAL)

r = add_sec(ws1, r, MID_BLUE, "VISIT SEATTLE  (Official DMO — Seattle/King County)", NC1)
vs_rows = [
    ["Visit Seattle", "Tourism — Destination Dev",
     "Marco Leal", "VP, Destination Development",
     "mleal@visitseattle.org", "206.461.5816",
     "Hi Marco — I'm launching Pierway, an AI-powered walking platform that converts pedestrian exploration into verified local commerce for Seattle merchants. Given your destination development mandate, I think there's a strong alignment. Would love 20 minutes.",
     "visitseattle.org/about-us/leadership/"],
    ["Visit Seattle", "Tourism — Executive",
     "Tammy Canavan", "President & CEO",
     "tcanavan@visitseattle.org", "206.461.5833",
     "Hi Tammy — Pierway is building AI walking routes for the 2M+ cruise passengers arriving in Seattle in 2026. We're looking to partner with Visit Seattle to embed QR codes at terminals and hotels. Happy to share a one-pager at your convenience.",
     "visitseattle.org/about-us/leadership/"],
    ["Visit Seattle", "Tourism — Sales & Marketing",
     "Kelly Saling", "EVP Sales & Marketing & Chief Business Officer",
     "ksaling@visitseattle.org", "206.461.5802",
     "Hi Kelly — Pierway is a new AI walking experience platform targeting Seattle's 2026 cruise season. We see strong co-marketing potential with Visit Seattle — happy to explore what a partnership could look like.",
     "visitseattle.org/about-us/leadership/"],
    ["Visit Seattle", "Tourism — Marketing",
     "Stephanie Byington", "Chief Marketing Officer",
     "sbyington@visitseattle.org", "206.461.5809",
     "Hi Stephanie — We're building Pierway, a narrative-driven AI walking platform for Seattle cruise passengers, and documenting the build publicly. Potential co-marketing story around authentic Seattle discovery.",
     "visitseattle.org/about-us/leadership/"],
]
for row_data in vs_rows:
    r = add_row(ws1, r, row_data, LIGHT_BLUE)

r = add_sec(ws1, r, PURPLE, "STATE OF WASHINGTON TOURISM  (Statewide DMO)", NC1)
wa_rows = [
    ["State of WA Tourism", "Tourism — Partnerships",
     "Mike Moe", "Director, Strategic Partnerships & Tourism Development",
     "tourisminfo@stateofwatourism.com", "",
     "Hi Mike — Pierway is a new AI walking experience platform launching at the Port of Seattle for the 2026 cruise season. Given your statewide partnerships mandate, I'd love to explore whether Pierway fits into WA Tourism's distribution strategy.",
     "industry.stateofwatourism.com/staff-and-board/"],
    ["State of WA Tourism", "Tourism — Executive",
     "David Blandford", "CEO",
     "tourisminfo@stateofwatourism.com", "",
     "Hi David — Launching Pierway in Seattle for cruise season 2026 — an AI walking platform that drives verified foot traffic to local merchants. Given your background at Visit Seattle, I'd value your perspective on where it fits in the state tourism ecosystem.",
     "industry.stateofwatourism.com/staff-and-board/"],
]
for row_data in wa_rows:
    r = add_row(ws1, r, row_data, LIGHT_PURP)

widths(ws1, [22, 20, 22, 34, 32, 16, 60, 40, 18, 18, 18, 18, 28])
ws1.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — CRUISE LINES
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Cruise Lines")

H2 = (["Company", "Terminal", "Priority", "New 2026?",
        "Contact Name", "Title", "Email", "Phone",
        "Targeted Pitch (QR / Pierway)", "Source / LinkedIn"] + TRACK_HDRS)
ws2.append(H2)
hdr(ws2, 1, DARK_BLUE)
NC2 = len(H2)

cruise_rows = [
    ("_SEC", "HOLLAND AMERICA LINE  |  hollandamerica.com  |  Pier 91  |  ~350,000+ pax  |  PRIORITY: HIGH", TEAL),
    ["Holland America Line", "Pier 91", "HIGH", "No",
     "Jessica Ashe", "Sr. Director, Shore Excursions & Future Cruises",
     "", "",
     "Hi Jessica — I'm launching Pierway, an AI walking experience platform for cruise passengers disembarking in Seattle. I'd love to explore embedding Pierway QR codes in HAL's shore excursion materials as a complimentary digital option for passengers with self-guided time. Could we connect?",
     "linkedin.com/in/jessicaashe55/"],
    ["Holland America Line", "Pier 91", "HIGH", "No",
     "Carole Biencourt", "VP, Onboard Revenue",
     "", "1-888-425-9376",
     "Hi Carole — Following your January 2026 announcement on new cultural tours, I wanted to introduce Pierway — an AI walking platform for Seattle port days that complements HAL's shore excursion offering with a self-guided narrative option.",
     "prnewswire.com — HAL 150 Cultural Tours Jan 2026"],
    ["Holland America Line", "Pier 91", "HIGH", "No",
     "Shore Excursions (dept)", "Shore Excursions Department",
     "hal_shore_excursions@hollandamerica.com", "206-626-7320",
     "Public dept email — use for cold outreach if named contacts don't respond. Reference Seattle port day experience for passengers.",
     "hollandamerica.com/contact-us"],
    ["Holland America Line", "Pier 91", "HIGH", "No",
     "Partnerships (general)", "Partnerships Inbox",
     "partnerships@hollandamerica.com", "",
     "Public partnerships email. HAL HQ is at 450 Third Ave W, Seattle WA 98119 — same city as Pierway.",
     "hollandamerica.com/contact-us"],

    ("_SEC", "PRINCESS CRUISES  |  princess.com  |  Pier 91  |  ~300,000+ pax  |  PRIORITY: HIGH", TEAL),
    ["Princess Cruises", "Pier 91", "HIGH", "No",
     "Wilkin Mes", "VP, Port Operations",
     "", "1-800-774-6237",
     "Hi Wilkin — I'm launching Pierway, an AI walking platform for cruise passengers during Seattle port days. I'd love to understand how Princess manages vendor placement at Pier 91, and whether a QR code partnership would fit within your port operations framework.",
     "princess.com/news/news-releases/2023/06/cruise-and-maritime-veteran-wilkin-mes"],
    ["Princess Cruises", "Pier 91", "HIGH", "No",
     "Terry Thornton", "Chief Commercial Officer",
     "", "1-800-774-6237",
     "Hi Terry — Pierway is a new AI-powered walking platform for Seattle port days that could complement Princess's shore excursion offering. Happy to share a one-pager on the commercial model — CPV-based, no upfront cost to Princess.",
     "princess.com/news/news-releases/2023/03/terry-thornton-named"],

    ("_SEC", "NORWEGIAN CRUISE LINE  |  ncl.com  |  Pier 66  |  ~250,000+ pax  |  PRIORITY: HIGH", TEAL),
    ["Norwegian Cruise Line", "Pier 66", "HIGH", "No",
     "Katty Byrd", "SVP, Guest Services (incl. Shore Excursions)",
     "kbyrd@ncl.com", "1-866-625-1164",
     "Hi Katty — I'm building Pierway, an AI walking experience for NCL passengers during Seattle port days. It's a zero-cost, self-guided complement to your existing shore excursion menu. Would love 20 minutes to show you what we're building.",
     "elliott.org/company-contacts/ncl | ncl.com/newsroom/katty-byrd"],
    ["Norwegian Cruise Line", "Pier 66", "HIGH", "No",
     "Milos Cicic", "Manager, Destination Services, Shore Excursions",
     "", "",
     "Hi Milos — Pierway is an AI-powered walking platform launching in Seattle for cruise passengers. I'd love to explore whether it fits within NCL's destination services offering as a complimentary self-guided option.",
     '"Norwegian Cruise Line" "Milos Cicic"'],

    ("_SEC", "VIRGIN VOYAGES  |  virginvoyages.com  |  Pier 91  |  NEW 2026 ★  |  PRIORITY: HIGHEST — outreach now", ORANGE),
    ["Virgin Voyages", "Pier 91", "HIGHEST ★", "YES — 2026",
     "Koreen McNutt", "VP, Agency & Business Development",
     "", "",
     "Hi Koreen — Congrats on the Seattle launch with Brilliant Lady! I'm building Pierway — an AI walking experience for Virgin passengers during Seattle port days. Given Voyages' ethos around curated experiences, I think Pierway's narrative-native approach could be a strong fit. Would love to connect before the season opens.",
     "linkedin.com/in/koreen-mcnutt/"],
    ["Virgin Voyages", "Pier 91", "HIGHEST ★", "YES — 2026",
     "Kristy Woolums", "Sr. Director, National Strategic Accounts",
     "", "",
     "Hi Kristy — I'm launching Pierway in Seattle for the 2026 Alaska season. Given your focus on national strategic partnerships, I'd love to explore whether Pierway's AI walking experience could be embedded as a value-add for Virgin Sailors during Seattle port days.",
     "travelpulse.com — VV Sales Appointments"],

    ("_SEC", "MSC CRUISES  |  msccruises.com  |  Pier 91  |  NEW 2026 ★  |  PRIORITY: HIGHEST — outreach now", ORANGE),
    ["MSC Cruises", "Pier 91", "HIGHEST ★", "YES — 2026",
     "Lynn Torrent", "President, MSC Cruises North America",
     "", "",
     "Hi Lynn — Exciting to see MSC bringing MSC Poesia to Seattle in 2026. I'm building Pierway — an AI-powered walking experience for cruise passengers during port days. As MSC stands up its Seattle program, I'd love to explore a QR placement partnership early. Happy to send a one-pager.",
     "linkedin.com/in/lynn-torrent-a2a0ab119/"],
    ["MSC Cruises", "Pier 91", "HIGHEST ★", "YES — 2026",
     "Gianluca Suprani", "SVP, Global Port Development & Shore Activities",
     "", "",
     "Hi Gianluca — I'm building Pierway, an AI walking experience platform launching at the Port of Seattle for the 2026 season. As the person overseeing global port development and shore activities, I'd love your perspective on how Pierway could complement MSC's port day offering.",
     "seatrade-cruise.com/people-opinions — Gianluca Suprani"],

    ("_SEC", "MEDIUM PRIORITY  |  Royal Caribbean / Celebrity / Carnival", MID_BLUE),
    ["Royal Caribbean International", "Pier 91", "MEDIUM", "No",
     "", "Destination Experiences Team",
     "", "",
     "Pierway pitch: AI walking routes for Seattle port days, performance-based model, QR code placement, complements shore excursion menu.",
     '"Royal Caribbean" "Destination Experiences" Seattle'],
    ["Celebrity Cruises", "Pier 91", "MEDIUM", "No",
     "", "Shore Excursions / Partnerships",
     "", "",
     "Same pitch as Royal Caribbean. Same parent (RCG) — one contact may cover both.",
     '"Celebrity Cruises" "Shore Excursions" Seattle'],
    ["Carnival Cruise Line", "Pier 91", "MEDIUM", "No",
     "", "Shore Excursions / Destination",
     "", "",
     "Pierway pitch: AI walking platform, QR at disembarkation, performance-based CPV model.",
     '"Carnival Cruise Line" "Shore Excursions" Alaska Seattle'],
]

row_num = 2
alt = False
for item in cruise_rows:
    if isinstance(item, tuple) and item[0] == "_SEC":
        section_hdr(ws2, row_num, item[2], item[1], NC2)
        row_num += 1
        alt = False
    else:
        ws2.append(item + BLANK5)
        row_style(ws2, row_num, bg=LIGHT_GREY if alt else WHITE)
        row_num += 1
        alt = not alt

widths(ws2, [24, 12, 12, 10, 24, 32, 36, 16, 68, 44, 18, 18, 18, 18, 28])
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — HOTELS
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Hotels")

H3 = (["Hotel", "Terminal", "Distance", "Stars",
        "Contact Name", "Title", "Email", "Phone",
        "Targeted Pitch (QR / Pierway)", "Source / LinkedIn"] + TRACK_HDRS)
ws3.append(H3)
hdr(ws3, 1, MID_BLUE)
NC3 = len(H3)

hotel_rows = [
    ("_SEC", "PIER 66 — BELL STREET  (Downtown / Belltown — walkable from terminal)", TEAL),
    ["The Edgewater Hotel", "Pier 66", "0.14 mi — walk-in", "4★",
     "Ian McClendon", "General Manager",
     "", "(206) 792-5959",
     "Hi Ian — I'm launching Pierway, an AI walking experience for cruise passengers during Seattle port days. Given Edgewater's location literally steps from Pier 66, a QR code in your lobby or concierge desk would reach passengers the moment they're deciding what to do. No cost to the hotel — we'd love to discuss a partnership.",
     "Luxury Travel Magazine — Edgewater Refresh 2025"],
    ["Seattle Marriott Waterfront", "Pier 66", "0.1 mi", "4★",
     "Haley Connors", "Destination Sales Executive",
     "", "(206) 443-5000",
     "Hi Haley — I'm building Pierway, an AI-powered walking platform for cruise passengers. Given your role in destination sales and Marriott's Park & Cruise package, I think Pierway QR codes in guest rooms or the concierge area would add real value. Happy to explore a partnership.",
     "rocketreach.co — Seattle Marriott Waterfront"],
    ["Seattle Marriott Waterfront", "Pier 66", "0.1 mi", "4★",
     "Amrit Sandhu", "General Manager",
     "", "(206) 443-5000",
     "Hi Amrit — I'm launching Pierway for the 2026 cruise season — an AI walking experience that adds value to Seattle port days. Given Marriott Waterfront's position at Pier 66 and your existing cruise packages, this could be a natural partnership.",
     "rocketreach.co — Seattle Marriott Waterfront"],
    ["Mayflower Park Hotel", "Pier 66 & 91", "1.0 mi / 3.2 mi", "3★",
     "Leslie Womack", "Director of Sales",
     "", "(206) 382-6991",
     "Hi Leslie — Pierway is an AI walking platform launching for Seattle's 2026 cruise season. Given Mayflower's active cruise packages and shuttle to both terminals, your guests are exactly our target user. I'd love to discuss placing Pierway QR codes at the front desk or in cruise packages.",
     "rocketreach.co/leslie-womack"],
    ["Mayflower Park Hotel", "Pier 66 & 91", "1.0 mi / 3.2 mi", "3★",
     "Andrew Harris", "General Manager",
     "press@azulhospitality.com", "(206) 623-8700",
     "Hi Andrew — Congratulations on your appointment at Mayflower. I'm building Pierway — an AI walking experience for cruise passengers — and I'd love to explore embedding Pierway in Mayflower's cruise offering given your shuttle program to both Pier 66 and Pier 91.",
     "hotel-online.com — Andrew Harris Appointment Jan 2026"],
    ["The Sound Hotel", "Pier 66", "0.3 mi", "3★",
     "Kelly Keith", "Sales",
     "", "(206) 441-7456",
     "Hi — I'm building Pierway, an AI walking experience for cruise passengers in Seattle. The Sound Hotel's Belltown location and cruise transfer partnership make it a natural fit for Pierway QR code placement. Happy to connect with whoever leads cruise partnerships.",
     "linkedin.com/in/kelly-keith-a2585833/"],

    ("_SEC", "PIER 91 — SMITH COVE  (Queen Anne / Magnolia area)", MID_BLUE),
    ["Mediterranean Inn", "Pier 91", "1.5 mi", "3★",
     "Sheila Ordonez", "General Manager",
     "", "(206) 428-4700",
     "Hi Sheila — I'm launching Pierway, an AI walking experience for cruise passengers during Seattle port days. The Mediterranean Inn's proximity to Pier 91 and your active shuttle program make this a strong fit for a QR code partnership. No cost to the hotel — just a better guest experience.",
     "linkedin.com/in/sheilaordonez"],
    ["Astra Hotel Seattle", "Pier 91", "2.5 mi", "4★",
     "", "General Manager / Director of Sales",
     "", "",
     "Pierway pitch: Your 'Snooze and Cruise' package guests are exactly our target. QR code in room or at checkout could point them to Pierway for their port day experience.",
     '"Astra Hotel Seattle" "General Manager" OR "Director of Sales"'],
    ["Staypineapple The Maxwell Hotel", "Pier 91", "1.2 mi", "3★",
     "", "General Manager / Director of Sales",
     "", "",
     "Pierway pitch: Closest full-service hotel to Pier 91 with no existing cruise program — opportunity to be the first to offer a curated port day experience through Pierway QR codes.",
     '"Maxwell Hotel Seattle" OR "Staypineapple Seattle" "General Manager"'],
]

row_num = 2
alt = False
for item in hotel_rows:
    if isinstance(item, tuple) and item[0] == "_SEC":
        section_hdr(ws3, row_num, item[2], item[1], NC3)
        row_num += 1
        alt = False
    else:
        ws3.append(item + BLANK5)
        row_style(ws3, row_num, bg=LIGHT_GREY if alt else WHITE)
        row_num += 1
        alt = not alt

widths(ws3, [26, 12, 16, 6, 22, 28, 30, 16, 68, 44, 18, 18, 18, 18, 28])
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = ws3.dimensions


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — OTHER TARGETS
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Other Targets")

H4 = (["Organization", "Category", "Why They Matter",
        "Contact Name / Role to Find", "Website / Contact",
        "Targeted Pitch (QR / Pierway)"] + TRACK_HDRS)
ws4.append(H4)
hdr(ws4, 1, PURPLE)
NC4 = len(H4)

other_rows = [
    ("_SEC", "WATER-BASED TRANSPORT  (passengers already in transit mode)", TEAL),
    ["Argosy Cruises", "Local boat tours & ferries",
     "Kevin Clark (CEO) sits on Visit Seattle board — direct connection to tourism ecosystem. Argosy runs Seattle waterfront tours and charters. Tourist-facing, high foot traffic.",
     "Kevin Clark — CEO  |  Find GM / Partnerships Director",
     "argosycruises.com",
     "Hi — I'm building Pierway, an AI walking experience for visitors in Seattle. Argosy guests are exactly our target — arriving at the waterfront with time to explore on foot. A QR code on your boats or dock could funnel guests directly into Pierway routes."],
    ["Washington State Ferries", "Ferry system — Colman Dock",
     "Colman Dock (downtown Seattle) serves millions of ferry passengers annually. Foot passengers disembark directly into downtown Seattle — identical behavioral profile to cruise passengers.",
     "Find: Director of Passenger Experience or Marketing",
     "wsdot.wa.gov/ferries",
     "Pierway pitch: Ferry foot passengers disembark at Colman Dock with time to explore Seattle. QR code at the dock or on the ferry itself reaches a captive audience in exploration mode."],

    ("_SEC", "HIGH-TRAFFIC TOURIST ATTRACTIONS  (QR placement in physical space)", MID_BLUE),
    ["Pike Place Market", "Landmark / Attraction",
     "~10M visitors/year, mostly pedestrian, many first-timers. Visitor info center is a natural QR placement point.",
     "Find: Director of Marketing or Visitor Services Manager",
     "pikeplacemarket.org",
     "Pierway pitch: Pike Place is often the first stop for cruise passengers. A Pierway QR code at the visitor info center or market entrance puts us in front of pedestrians the moment they start exploring."],
    ["Space Needle / Seattle Center", "Landmark / Attraction",
     "Top tourist destination near Pier 91 corridor. Visitors are already in 'explore Seattle' mode.",
     "Find: Director of Guest Experience or Marketing",
     "spaceneedle.com | seattlecenter.com",
     "Pierway pitch: Space Needle visitors are tourists in active exploration mode. A QR code at ticketing or the base of the needle reaches the exact user Pierway is built for."],
    ["Chihuly Garden and Glass", "Landmark / Attraction",
     "Adjacent to Space Needle. High-value tourist attraction. Visitors are pre-qualified as engaged, curious tourists.",
     "Find: GM or Director of Marketing",
     "chihulygardenandglass.com",
     "Pierway pitch: Post-exhibit, Chihuly visitors are walking back into Seattle Center with nowhere to go next. Pierway fills that gap with a curated next stop."],
    ["Museum of Pop Culture (MoPOP)", "Museum / Attraction",
     "Seattle Center. Tourist-facing. High foot traffic from cruise visitors.",
     "Find: Director of Visitor Experience or Partnerships",
     "mopop.org",
     "Standard Pierway QR pitch — curated walking routes for visitors leaving the museum."],

    ("_SEC", "TOUR OPERATORS & EXCURSION COMPANIES  (existing shore excursion vendors)", ORANGE),
    ["Viator / TripAdvisor Experiences", "Online tour marketplace",
     "Viator is the dominant shore excursion marketplace used by cruise passengers pre-trip. Getting listed means Pierway appears when passengers search 'Seattle port day activities.'",
     "Find: Supplier Partnerships or Supplier Onboarding contact",
     "viator.com/partner",
     "Pierway listing pitch: Pierway as a free/freemium digital experience with paid merchant layer. List as a 'Self-Guided AI Walking Tour' category."],
    ["GetYourGuide", "Online tour marketplace",
     "Second-largest tours/activities marketplace globally. Used heavily by European cruise passengers (high overlap with MSC, Cunard, Silversea demographics).",
     "Find: Supplier Partnerships",
     "getyourguide.com/supplier",
     "Same as Viator — list Pierway as AI walking tour. European tourist focus aligns with MSC's Seattle debut."],
    ["Gray Line Seattle / Grayline Tours", "Local tour operator",
     "Established shore excursion vendor for Seattle cruise terminals. Existing relationship with cruise lines. Could white-label or co-distribute Pierway.",
     "Find: Operations Director or GM",
     "graylineseattle.com",
     "Pierway partnership pitch: We augment your bus tour offering with a self-guided AI option for passengers who prefer walking. Rev share model — you distribute, we pay per visit."],

    ("_SEC", "AIRPORT & PRE-ARRIVAL  (catching tourists before they reach the port)", PURPLE),
    ["Sea-Tac Airport — Ground Transportation / Concierge", "Airport",
     "Many cruise passengers fly into Sea-Tac the day before and go downtown. Airport concierge / info kiosks are a pre-arrival distribution point.",
     "Find: Director of Guest Experience, Airport Concierge Services, or Advertising/Retail Partnerships",
     "portseattle.org/sea-tac",
     "Pierway pitch: Catch arriving cruise passengers at baggage claim or the airport info center before they reach their hotel. A QR code here plants the seed a day early."],
    ["Alaska Airlines (Seattle Hub)", "Airline",
     "Alaska Airlines is Sea-Tac's dominant carrier and Seattle's home airline. Many Alaska cruise passengers fly Alaska Air. In-flight or lounge QR placement reaches them pre-arrival.",
     "Find: Director of Partnerships or Inflight Experience",
     "alaskaair.com",
     "Pierway pitch: Include Pierway in Alaska Airlines' Seattle destination content — seatback screens, Sky Magazine, or Lounge digital displays. Reach cruise passengers before they land."],
]

row_num = 2
alt = False
for item in other_rows:
    if isinstance(item, tuple) and item[0] == "_SEC":
        section_hdr(ws4, row_num, item[2], item[1], NC4)
        row_num += 1
        alt = False
    else:
        ws4.append(item + BLANK5)
        row_style(ws4, row_num, bg=LIGHT_GREY if alt else WHITE)
        row_num += 1
        alt = not alt

widths(ws4, [26, 20, 44, 30, 28, 68, 18, 18, 18, 18, 28])
ws4.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — HOW TO USE
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("How To Use")
ws5.column_dimensions["A"].width = 115

guide = [
    ("PIERWAY — Prospect & Outreach Tracker  |  Port of Seattle  |  v4  |  2026-03-12", True, DARK_BLUE, WHITE, 14),
    ("", False, WHITE, "000000", 10),
    ("TRACKING COLUMNS — HOW TO USE", True, TEAL, WHITE, 11),
    ("'Outreach Sent?' — Enter Y or the date you sent the message (e.g. Mar 14)", False, LIGHT_TEAL, "000000", 10),
    ("'Response?' — Enter Y or the date you received a reply", False, LIGHT_TEAL, "000000", 10),
    ("'Follow-up Needed?' — Y/N. Flag if no response after 5 business days", False, LIGHT_TEAL, "000000", 10),
    ("'Meeting Scheduled?' — Enter date of meeting if booked", False, LIGHT_TEAL, "000000", 10),
    ("'Status Notes' — Free text: outcome, next step, who else to contact, etc.", False, LIGHT_TEAL, "000000", 10),
    ("", False, WHITE, "000000", 10),
    ("RECOMMENDED OUTREACH SEQUENCE", True, MID_BLUE, WHITE, 11),
    ("Wave 1 (This Week): Marie Ellingson (Port), Marco Leal (Visit Seattle), Koreen McNutt (Virgin Voyages)", False, LIGHT_BLUE, "000000", 10),
    ("Wave 2 (Week 2): Lynn Torrent (MSC), Jessica Ashe + Carole Biencourt (HAL), Katty Byrd (NCL - kbyrd@ncl.com)", False, LIGHT_BLUE, "000000", 10),
    ("Wave 3 (Week 3): Hotels — Haley Connors (Marriott), Ian McClendon (Edgewater), Leslie Womack (Mayflower)", False, LIGHT_BLUE, "000000", 10),
    ("Wave 4 (Week 4): Wilkin Mes (Princess), tourism marketplace listings (Viator, GetYourGuide), Argosy Cruises", False, LIGHT_BLUE, "000000", 10),
    ("", False, WHITE, "000000", 10),
    ("KEY PUBLIC EMAILS CONFIRMED", True, MID_BLUE, WHITE, 11),
    ("Port of Seattle — Marie Ellingson:  mellingson@portseattle.org  |  (206) 787-3529", False, LIGHT_GREY, "000000", 10),
    ("Port of Seattle — Cruise Terminal:  cruiseterminal@portseattle.org  |  (206) 644-1355", False, LIGHT_GREY, "000000", 10),
    ("HAL — Shore Excursions:             hal_shore_excursions@hollandamerica.com", False, LIGHT_GREY, "000000", 10),
    ("HAL — Partnerships:                 partnerships@hollandamerica.com", False, LIGHT_GREY, "000000", 10),
    ("NCL — Katty Byrd (SVP):             kbyrd@ncl.com  (Elliott Report public directory)", False, LIGHT_GREY, "000000", 10),
    ("Visit Seattle — Marco Leal:         mleal@visitseattle.org  |  206.461.5816", False, LIGHT_GREY, "000000", 10),
    ("Visit Seattle — Tammy Canavan:      tcanavan@visitseattle.org  |  206.461.5833", False, LIGHT_GREY, "000000", 10),
    ("Visit Seattle — Kelly Saling:       ksaling@visitseattle.org  |  206.461.5802", False, LIGHT_GREY, "000000", 10),
    ("State of WA Tourism:                tourisminfo@stateofwatourism.com", False, LIGHT_GREY, "000000", 10),
    ("Mayflower — Leslie Womack (direct): (206) 382-6991", False, LIGHT_GREY, "000000", 10),
    ("MSC Cruises NA — Lynn Torrent:      linkedin.com/in/lynn-torrent-a2a0ab119/", False, LIGHT_GREY, "000000", 10),
    ("", False, WHITE, "000000", 10),
    ("SEASON DEADLINE", True, MID_BLUE, WHITE, 11),
    ("Cruise season opens mid-April 2026. Target all distribution deals signed by April 1.", False, LIGHT_BLUE, "000000", 10),
    ("Virgin Voyages (Brilliant Lady) and MSC Cruises (MSC Poesia) both homeport at Pier 91 for the first time in 2026.", False, LIGHT_BLUE, "000000", 10),
    ("2026 season projected: 2M+ passengers. 298 ship calls. ~$1.2B economic impact.", False, LIGHT_BLUE, "000000", 10),
    ("", False, WHITE, "000000", 10),
    ("RESEARCH SOURCES", True, MID_BLUE, WHITE, 11),
    ("portseattle.org | visitseattle.org | industry.stateofwatourism.com | hollandamerica.com | ncl.com/newsroom", False, LIGHT_GREY, "000000", 10),
    ("elliott.org/company-contacts/ncl | princess.com/news | seatrade-cruise.com | travelmarketreport.com", False, LIGHT_GREY, "000000", 10),
    ("travelpulse.com | hotel-online.com | hospitalitynet.org | rocketreach.co | prnewswire.com", False, LIGHT_GREY, "000000", 10),
    ("Compiled: 2026-03-12 by Claude (Pierway project)", False, LIGHT_GREY, "000000", 10),
]

for text, bold, bg, fg, size in guide:
    ws5.append([text])
    rr = ws5.max_row
    c = ws5.cell(row=rr, column=1)
    c.font      = Font(bold=bold, color=fg, name="Calibri", size=size)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws5.row_dimensions[rr].height = 22 if bold else 20

out = r"C:\Users\raysc\wander\research\port-of-seattle-prospects.xlsx"
wb.save(out)
print(f"Saved: {out}")
